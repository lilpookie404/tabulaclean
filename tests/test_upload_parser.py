from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from server.uploads.errors import UploadError
from server.uploads.parser import ParseLimits, parse_upload


def _workbook_bytes(
    *,
    hidden_rows: list[list[object]] | None = None,
    visible_rows: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    hidden = workbook.active
    hidden.title = "Hidden"
    for row in hidden_rows or [["ignore"], ["old"]]:
        hidden.append(row)

    visible = workbook.create_sheet("Contacts")
    for row in visible_rows or [["name", "joined"], ["Aarav", "2026-01-02"]]:
        visible.append(row)

    hidden.sheet_state = "hidden"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_csv_preserves_duplicate_blank_headers_and_lexical_values() -> None:
    parsed = parse_upload(
        "contacts.csv",
        b"name,,name,amount\n Aarav ,x,Alias,0012.50\n",
    )

    assert parsed.display_headers == ["name", "", "name", "amount"]
    assert parsed.column_ids == ["column_1", "column_2", "column_3", "column_4"]
    assert parsed.dataframe.columns.tolist() == parsed.column_ids
    assert parsed.dataframe.iloc[0].to_dict() == {
        "column_1": " Aarav ",
        "column_2": "x",
        "column_3": "Alias",
        "column_4": "0012.50",
    }
    assert parsed.sheet_name is None


def test_parse_csv_accepts_utf8_bom_and_windows_1252() -> None:
    utf8 = parse_upload("utf8.csv", "\ufeffcity\nMontréal\n".encode())
    windows = parse_upload("windows.csv", "city\nMontréal\n".encode("cp1252"))

    assert utf8.display_headers == ["city"]
    assert utf8.dataframe.iloc[0, 0] == "Montréal"
    assert windows.dataframe.iloc[0, 0] == "Montréal"


def test_parse_csv_preserves_blank_rows_for_source_row_numbers() -> None:
    parsed = parse_upload(
        "contacts.csv",
        b"name,email\nAarav,a@example.com\n,\nMeera,m@example.com\n",
    )

    assert parsed.dataframe.shape == (3, 2)
    assert parsed.dataframe.iloc[1].tolist() == ["", ""]
    assert parsed.dataframe.iloc[2, 0] == "Meera"


def test_parse_xlsx_uses_first_visible_worksheet() -> None:
    parsed = parse_upload(
        "contacts.xlsx",
        _workbook_bytes(
            visible_rows=[
                ["customer", "joined"],
                ["Meera", "2026-03-14"],
                ["Riya", "2026-03-15"],
            ]
        ),
    )

    assert parsed.sheet_name == "Contacts"
    assert parsed.display_headers == ["customer", "joined"]
    assert parsed.dataframe.shape == (2, 2)
    assert parsed.dataframe.iloc[0].to_dict() == {
        "column_1": "Meera",
        "column_2": "2026-03-14",
    }


@pytest.mark.parametrize(
    ("filename", "payload"),
    [
        ("blank.csv", b"name,email\n,\n"),
        (
            "blank.xlsx",
            _workbook_bytes(visible_rows=[["name", "email"], [None, None]]),
        ),
    ],
)
def test_parse_upload_rejects_tables_without_populated_data_rows(
    filename: str,
    payload: bytes,
) -> None:
    with pytest.raises(UploadError) as captured:
        parse_upload(filename, payload)

    assert captured.value.code == "empty_table"
    assert captured.value.status_code == 422


@pytest.mark.parametrize(
    ("filename", "payload", "code", "status_code"),
    [
        ("contacts.txt", b"name\nAarav\n", "unsupported_file_type", 415),
        ("contacts.csv", b"", "empty_file", 422),
        ("contacts.csv", b"name\n", "empty_table", 422),
        ("contacts.csv", b"a,b\n1,2,3\n", "invalid_file", 422),
        ("contacts.xlsx", b"not a workbook", "invalid_file", 422),
    ],
)
def test_parse_upload_returns_clean_errors(
    filename: str,
    payload: bytes,
    code: str,
    status_code: int,
) -> None:
    with pytest.raises(UploadError) as captured:
        parse_upload(filename, payload)

    assert captured.value.code == code
    assert captured.value.status_code == status_code
    assert captured.value.message


def test_parse_upload_enforces_table_dimensions() -> None:
    with pytest.raises(UploadError) as rows_error:
        parse_upload(
            "rows.csv",
            b"name\nA\nB\n",
            limits=ParseLimits(max_rows=1),
        )
    with pytest.raises(UploadError) as columns_error:
        parse_upload(
            "columns.csv",
            b"a,b\n1,2\n",
            limits=ParseLimits(max_columns=1),
        )

    assert rows_error.value.code == "table_too_large"
    assert columns_error.value.code == "table_too_large"
    assert rows_error.value.status_code == 413


def test_parse_upload_enforces_dataframe_memory_limit() -> None:
    with pytest.raises(UploadError) as captured:
        parse_upload(
            "contacts.csv",
            b"name\nAarav\n",
            limits=ParseLimits(max_dataframe_bytes=1),
        )

    assert captured.value.code == "table_too_large"
    assert captured.value.status_code == 413


def test_parse_xlsx_rejects_suspicious_archive_expansion() -> None:
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.xml", "x" * 1_000)

    with pytest.raises(UploadError) as captured:
        parse_upload(
            "large.xlsx",
            buffer.getvalue(),
            limits=ParseLimits(max_archive_uncompressed_bytes=100),
        )

    assert captured.value.code == "file_too_large"
    assert captured.value.status_code == 413


def test_parse_xlsx_returns_friendly_error_for_malformed_workbook_xml() -> None:
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types>")
        archive.writestr("_rels/.rels", "<Relationships>")
        archive.writestr("xl/workbook.xml", "<workbook>")

    with pytest.raises(UploadError) as captured:
        parse_upload("malformed.xlsx", buffer.getvalue())

    assert captured.value.code == "invalid_file"
    assert captured.value.status_code == 422


def test_parse_xlsx_rejects_oversized_dimensions_before_reading_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class OversizedSheet:
        sheet_state = "visible"
        title = "Oversized"
        max_row = 3
        max_column = 1

        def iter_rows(self, *, values_only: bool):
            del values_only
            raise AssertionError("Rows should not be materialized after the dimension check.")

    class FakeWorkbook:
        worksheets = [OversizedSheet()]

        def close(self) -> None:
            pass

    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.xml", "<workbook />")

    monkeypatch.setattr(
        "server.uploads.parser.load_workbook",
        lambda *_args, **_kwargs: FakeWorkbook(),
    )

    with pytest.raises(UploadError) as captured:
        parse_upload(
            "oversized.xlsx",
            buffer.getvalue(),
            limits=ParseLimits(max_rows=1),
        )

    assert captured.value.code == "table_too_large"
    assert captured.value.status_code == 413
