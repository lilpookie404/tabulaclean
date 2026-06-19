"""Bounded CSV and XLSX parsing for temporary upload sessions."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .errors import UploadError


MEBIBYTE = 1024 * 1024


@dataclass(frozen=True)
class ParseLimits:
    max_upload_bytes: int = 10 * MEBIBYTE
    max_rows: int = 100_000
    max_columns: int = 200
    max_dataframe_bytes: int = 100 * MEBIBYTE
    max_archive_uncompressed_bytes: int = 100 * MEBIBYTE

    @property
    def max_upload_mebibytes(self) -> int:
        return max(1, self.max_upload_bytes // MEBIBYTE)


@dataclass
class ParsedTable:
    filename: str
    sheet_name: str | None
    display_headers: list[str]
    column_ids: list[str]
    dataframe: pd.DataFrame
    memory_bytes: int


def _invalid_file(message: str) -> UploadError:
    return UploadError(422, "invalid_file", message)


def upload_size_error_message(limits: ParseLimits) -> str:
    return (
        "Please choose a spreadsheet no larger than "
        f"{limits.max_upload_mebibytes} MB."
    )


def _display_header(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value)


def _is_populated(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    try:
        return not bool(pd.isna(value))
    except (TypeError, ValueError):
        return True


def _build_table(
    *,
    filename: str,
    sheet_name: str | None,
    headers: list[Any],
    rows: list[list[Any]] | pd.DataFrame,
    limits: ParseLimits,
) -> ParsedTable:
    display_headers = [_display_header(value) for value in headers]
    if not display_headers:
        raise UploadError(422, "empty_table", "The spreadsheet does not contain any columns.")

    column_ids = [f"column_{position + 1}" for position in range(len(display_headers))]
    dataframe = (
        rows.copy()
        if isinstance(rows, pd.DataFrame)
        else pd.DataFrame(rows, columns=column_ids)
    )
    dataframe.columns = column_ids
    dataframe.reset_index(drop=True, inplace=True)

    if dataframe.empty:
        raise UploadError(422, "empty_table", "The spreadsheet does not contain any data rows.")
    if not any(
        _is_populated(value)
        for row in dataframe.itertuples(index=False, name=None)
        for value in row
    ):
        raise UploadError(422, "empty_table", "The spreadsheet does not contain any data rows.")
    if len(dataframe) > limits.max_rows or len(dataframe.columns) > limits.max_columns:
        raise UploadError(
            413,
            "table_too_large",
            (
                f"Please use a table with at most {limits.max_rows:,} rows "
                f"and {limits.max_columns:,} columns."
            ),
        )

    memory_bytes = int(dataframe.memory_usage(index=True, deep=True).sum())
    if memory_bytes > limits.max_dataframe_bytes:
        raise UploadError(
            413,
            "table_too_large",
            "This spreadsheet expands to more data than a temporary session can safely hold.",
        )

    return ParsedTable(
        filename=filename,
        sheet_name=sheet_name,
        display_headers=display_headers,
        column_ids=column_ids,
        dataframe=dataframe,
        memory_bytes=memory_bytes,
    )


def _parse_csv(filename: str, payload: bytes, limits: ParseLimits) -> ParsedTable:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = payload.decode("cp1252")
        except UnicodeDecodeError as exc:  # pragma: no cover - cp1252 maps every byte
            raise _invalid_file("The CSV text encoding could not be read.") from exc

    try:
        raw = pd.read_csv(
            StringIO(text),
            header=None,
            dtype=str,
            keep_default_na=False,
            na_filter=False,
            sep=",",
            engine="python",
            on_bad_lines="error",
            skip_blank_lines=False,
        )
    except (pd.errors.ParserError, UnicodeError, ValueError) as exc:
        raise _invalid_file("The CSV file is malformed or could not be read.") from exc

    if raw.empty:
        raise UploadError(422, "empty_table", "The spreadsheet does not contain any data rows.")

    headers = raw.iloc[0].tolist()
    data = raw.iloc[1:].copy()
    return _build_table(
        filename=filename,
        sheet_name=None,
        headers=headers,
        rows=data,
        limits=limits,
    )


def _inspect_xlsx(payload: bytes, limits: ParseLimits) -> None:
    try:
        with ZipFile(BytesIO(payload)) as archive:
            names = set(archive.namelist())
            if {"EncryptionInfo", "EncryptedPackage"} & names:
                raise _invalid_file("Encrypted Excel workbooks are not supported.")
            expanded_bytes = sum(member.file_size for member in archive.infolist())
            if expanded_bytes > limits.max_archive_uncompressed_bytes:
                raise UploadError(
                    413,
                    "file_too_large",
                    "This Excel workbook expands beyond the safe processing limit.",
                )
    except BadZipFile as exc:
        raise _invalid_file("The Excel workbook is corrupt, encrypted, or unreadable.") from exc


def _parse_xlsx(filename: str, payload: bytes, limits: ParseLimits) -> ParsedTable:
    _inspect_xlsx(payload, limits)
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except (
        BadZipFile,
        EOFError,
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        ValueError,
    ) as exc:
        raise _invalid_file("The Excel workbook is corrupt, encrypted, or unreadable.") from exc

    try:
        visible_sheet = next(
            (sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"),
            None,
        )
        if visible_sheet is None:
            raise _invalid_file("The Excel workbook does not contain a visible worksheet.")

        if (
            max((visible_sheet.max_row or 0) - 1, 0) > limits.max_rows
            or (visible_sheet.max_column or 0) > limits.max_columns
        ):
            raise UploadError(
                413,
                "table_too_large",
                (
                    f"Please use a table with at most {limits.max_rows:,} rows "
                    f"and {limits.max_columns:,} columns."
                ),
            )

        row_iterator = visible_sheet.iter_rows(values_only=True)
        try:
            headers = list(next(row_iterator))
        except StopIteration:
            raise UploadError(422, "empty_table", "The spreadsheet does not contain any data rows.")

        rows: list[list[Any]] = []
        for row_number, row in enumerate(row_iterator, start=1):
            if row_number > limits.max_rows or len(row) > limits.max_columns:
                raise UploadError(
                    413,
                    "table_too_large",
                    (
                        f"Please use a table with at most {limits.max_rows:,} rows "
                        f"and {limits.max_columns:,} columns."
                    ),
                )
            rows.append(list(row))

        return _build_table(
            filename=filename,
            sheet_name=visible_sheet.title,
            headers=headers,
            rows=rows,
            limits=limits,
        )
    finally:
        workbook.close()


def parse_upload(
    filename: str,
    payload: bytes,
    *,
    limits: ParseLimits | None = None,
) -> ParsedTable:
    selected_limits = limits or ParseLimits()
    suffix = Path(filename).suffix.lower()

    if suffix not in {".csv", ".xlsx"}:
        raise UploadError(
            415,
            "unsupported_file_type",
            "Please choose a CSV or XLSX spreadsheet.",
        )
    if not payload:
        raise UploadError(422, "empty_file", "The selected file is empty.")
    if len(payload) > selected_limits.max_upload_bytes:
        raise UploadError(
            413,
            "file_too_large",
            upload_size_error_message(selected_limits),
        )

    if suffix == ".csv":
        return _parse_csv(filename, payload, selected_limits)
    return _parse_xlsx(filename, payload, selected_limits)
